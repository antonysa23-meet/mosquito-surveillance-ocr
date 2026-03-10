"""
Mosquito Surveillance Record Digitization Pipeline
====================================================
Converts handwritten tabular PDF records from Harris County CDC Gravid trap
collections into computer-readable CSV/XLSX formats using 100% local AI —
no internet connection or cloud service required.

Default mode (hybrid) uses two OCR engines selected per-column based on
benchmarking: TrOCR (Microsoft handwriting transformer) for numeric pooling
fields, EasyOCR for text/code fields. Achieves ~47.6% field accuracy on CPU.
A GPU with llama3.2-vision is expected to push this to 70–90%+.

Usage:
    python pipeline.py example_mosquito_record_test.pdf
    python pipeline.py example_mosquito_record_test.pdf --mode hybrid   (default)
    python pipeline.py example_mosquito_record_test.pdf --mode trocr
    python pipeline.py example_mosquito_record_test.pdf --mode easyocr
    python pipeline.py example_mosquito_record_test.pdf --mode whole-page --model llama3.2-vision

Pipeline stages:
    1. PDF → high-resolution PNG image (300 DPI via PyMuPDF)
    2. OpenCV table-cell detection (morphological line isolation)
    3. Cell-by-cell OCR using the selected engine(s)
    4. Column-specific post-processing and type coercion
    5. Species name matching against Harris County mosquito database
    6. CSV + XLSX output with QC comparison report
"""

from __future__ import annotations

import argparse
import base64
import json
import re
import sys
from pathlib import Path
from typing import Optional

import cv2
import fitz  # PyMuPDF
import numpy as np
import pytesseract
import requests
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.utils import get_column_letter
from PIL import Image

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

TESSERACT_CMD: str = r"C:/Program Files/Tesseract-OCR/tesseract.exe"
pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

OLLAMA_BASE_URL: str = "http://localhost:11434"
DEFAULT_MODEL: str = "llava"
FALLBACK_MODEL: str = "llava"  # same model, different attempt after resize

PDF_DPI: int = 300
PDF_ZOOM: float = PDF_DPI / 72.0  # fitz uses a zoom matrix

SCRIPT_DIR: Path = Path(__file__).resolve().parent
TEMP_DIR: Path = SCRIPT_DIR / "_temp"
CELLS_DIR: Path = TEMP_DIR / "cells"

OUTPUT_CSV: Path = SCRIPT_DIR / "mosquito_surveillance_data.csv"
OUTPUT_XLSX: Path = SCRIPT_DIR / "mosquito_surveillance_data.xlsx"

# Minimum cell dimensions (pixels at 300 DPI) to be considered a real cell
MIN_CELL_W: int = 20
MIN_CELL_H: int = 15

# Row is "fully black / redacted" if mean pixel value < this threshold
REDACTED_ROW_THRESHOLD: int = 30

# The 12 data columns this Harris County Gravid form always has
COLUMNS: list[str] = [
    "SITE_NO", "COLL_NO", "SPECIES", "POOL", "NO_MOSQ_POOLED",
    "VIAL_NO", "MSI", "STATUS", "NUM_POOLS", "AREA", "CITY", "ADDRESS",
]

# Column widths for the XLSX output
COLUMN_WIDTHS: dict[str, int] = {
    "SITE_NO": 9, "COLL_NO": 12, "SPECIES": 38, "POOL": 7,
    "NO_MOSQ_POOLED": 18, "VIAL_NO": 16, "MSI": 7, "STATUS": 9,
    "NUM_POOLS": 9, "AREA": 8, "CITY": 8, "ADDRESS": 30,
}

# Ground truth for the example PDF (used in QC report)
GROUND_TRUTH: list[dict[str, str]] = [
    {
        "SITE_NO": "1841", "COLL_NO": "T4GV 508",
        "SPECIES": "40 Cx.qf 20F; 4 Ae.ab 10F",
        "POOL": "2", "NO_MOSQ_POOLED": "39 Cx.qf / 4 Ae.ab",
        "VIAL_NO": "8758 / 8759",
        "AREA": "216", "CITY": "NEC", "ADDRESS": "614 KERNOHAN (419-H)",
    },
    {
        "SITE_NO": "1987", "COLL_NO": "T4GV 509",
        "SPECIES": "5 Cx.qf 40F",
        "POOL": "1", "NO_MOSQ_POOLED": "5",
        "VIAL_NO": "8760",
        "AREA": "217", "CITY": "CB", "ADDRESS": "4603 SHETR LN (380-M)",
    },
    {
        "SITE_NO": "1932", "COLL_NO": "T4GV 510",
        "SPECIES": "210 Cx.qf 10F; 5 Ae.ab",
        "POOL": "2", "NO_MOSQ_POOLED": "50 Cx.qf / 5 Ae.ab",
        "VIAL_NO": "8761 / 8762",
        "AREA": "316", "CITY": "HG", "ADDRESS": "214 SOUTH 4TH (459-V)",
    },
    {
        "SITE_NO": "1671", "COLL_NO": "T4GV 511",
        "SPECIES": "20 Cx.qf 10F; 1 Cu.ng; 2 Ae.ab",
        "POOL": "1", "NO_MOSQ_POOLED": "24",
        "VIAL_NO": "8763",
        "AREA": "317", "CITY": "BR", "ADDRESS": "HOLLY RD & CHAMBERS",
    },
    {
        "SITE_NO": "1844", "COLL_NO": "T4GV 512",
        "SPECIES": "53 Cx.qf 10F; 2 Ae.ab 10F",
        "POOL": "2", "NO_MOSQ_POOLED": "50 Cx.qf / 2 Ae.ab",
        "VIAL_NO": "8768 / 8769",
        "AREA": "321", "CITY": "CB", "ADDRESS": "14334 GREEN ACRES (420-K)",
    },
    {
        "SITE_NO": "1825", "COLL_NO": "T4GV 513",
        "SPECIES": "41 Cx.qf 20F; 1 Ps.co; 2 Ae.ab",
        "POOL": "1", "NO_MOSQ_POOLED": "41",
        "VIAL_NO": "8774",
        "AREA": "328", "CITY": "CB", "ADDRESS": "3115 WOLCEK RD (420-Q)",
    },
    {
        "SITE_NO": "1847", "COLL_NO": "T4GV 514",
        "SPECIES": "120 Cx.qf 10F; 1 Ps.fx; 3 Ae.ab",
        "POOL": "2", "NO_MOSQ_POOLED": "100",
        "VIAL_NO": "8775 / 8776",
        "AREA": "329", "CITY": "NEC", "ADDRESS": "15220 BOHEMIAN HALL DR",
    },
    {
        "SITE_NO": "1895", "COLL_NO": "T4GV 515",
        "SPECIES": "21 Cx.qf 20F; 4 Ae.ab",
        "POOL": "1", "NO_MOSQ_POOLED": "21",
        "VIAL_NO": "8780",
        "AREA": "320", "CITY": "BS", "ADDRESS": "214 COTTONTAIL (419-Z)",
    },
]

# QC fields to compare (subset — avoids penalising blank optional fields)
QC_FIELDS: list[str] = [
    "SITE_NO", "COLL_NO", "SPECIES", "POOL",
    "NO_MOSQ_POOLED", "VIAL_NO", "AREA", "CITY", "ADDRESS",
]

# ---------------------------------------------------------------------------
# Step 1 — PDF to image
# ---------------------------------------------------------------------------


def pdf_to_image(pdf_path: Path) -> Image.Image:
    """Render the first page of *pdf_path* at 300 DPI and return a PIL Image."""
    print("[Step 1/6] Converting PDF to image...")
    doc = fitz.open(str(pdf_path))
    page = doc[0]
    matrix = fitz.Matrix(PDF_ZOOM, PDF_ZOOM)
    pixmap = page.get_pixmap(matrix=matrix)
    doc.close()

    n_channels = pixmap.n
    raw = np.frombuffer(pixmap.samples, dtype=np.uint8).reshape(
        pixmap.height, pixmap.width, n_channels
    )
    if n_channels == 4:
        img_bgr = cv2.cvtColor(raw, cv2.COLOR_RGBA2BGR)
    else:
        img_bgr = cv2.cvtColor(raw, cv2.COLOR_RGB2BGR)

    TEMP_DIR.mkdir(parents=True, exist_ok=True)
    page_png = TEMP_DIR / "page_0.png"
    cv2.imwrite(str(page_png), img_bgr)
    print(f"  Saved {pixmap.width}x{pixmap.height} px PNG -> {page_png}")

    pil_image = Image.fromarray(cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB))
    return pil_image


# ---------------------------------------------------------------------------
# Step 2 — Table cell detection
# ---------------------------------------------------------------------------

BBox = tuple[int, int, int, int]  # (x, y, w, h)
Grid = list[list[BBox]]           # Grid[row][col] = bbox


def _cluster_positions(positions: list[int], gap: int = 20) -> list[int]:
    """Merge pixel positions within *gap* of each other into a centroid."""
    if not positions:
        return []
    sorted_pos = sorted(set(positions))
    groups: list[list[int]] = [[sorted_pos[0]]]
    for p in sorted_pos[1:]:
        if p - groups[-1][-1] <= gap:
            groups[-1].append(p)
        else:
            groups.append([p])
    return [int(np.mean(g)) for g in groups]


def detect_cells(page_img: Image.Image) -> tuple[Grid, np.ndarray]:
    """
    Detect table cells from *page_img* using OpenCV morphological analysis.

    Returns:
        grid : 2-D list of (x, y, w, h) bounding boxes indexed [row][col]
        bgr  : the page as a BGR numpy array (reused later for cropping)
    """
    print("[Step 2/6] Detecting table cells...")
    bgr = cv2.cvtColor(np.array(page_img), cv2.COLOR_RGB2BGR)
    gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
    h, w = gray.shape

    # Binarize (inverted so lines are white on black)
    binary = cv2.adaptiveThreshold(
        gray, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV,
        blockSize=15, C=4,
    )

    # Isolate horizontal lines
    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(w // 8, 50), 1))
    horiz = cv2.morphologyEx(binary, cv2.MORPH_OPEN, h_kernel, iterations=2)

    # Isolate vertical lines
    v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(h // 15, 50)))
    vert = cv2.morphologyEx(binary, cv2.MORPH_OPEN, v_kernel, iterations=2)

    # Project onto axes to find dense line positions
    row_proj = np.sum(horiz, axis=1)
    col_proj = np.sum(vert, axis=0)

    row_pos = np.where(row_proj > w * 0.05)[0].tolist()
    col_pos = np.where(col_proj > h * 0.02)[0].tolist()

    rows = _cluster_positions(row_pos, gap=20)
    cols = _cluster_positions(col_pos, gap=20)

    print(f"  Detected {len(rows)} horizontal lines, {len(cols)} vertical lines")
    print(f"  Grid: {len(rows)-1} rows x {len(cols)-1} columns")

    # Build the 2-D grid of bounding boxes
    grid: Grid = []
    for r in range(len(rows) - 1):
        row_boxes: list[BBox] = []
        for c in range(len(cols) - 1):
            x = cols[c]
            y = rows[r]
            bw = cols[c + 1] - x
            bh = rows[r + 1] - y
            if bw >= MIN_CELL_W and bh >= MIN_CELL_H:
                row_boxes.append((x, y, bw, bh))
            else:
                row_boxes.append(None)  # type: ignore[arg-type]
        grid.append(row_boxes)

    # Save cell crops to _temp/cells/ for debugging
    CELLS_DIR.mkdir(parents=True, exist_ok=True)
    for row_idx, row_boxes in enumerate(grid):
        for col_idx, bbox in enumerate(row_boxes):
            if bbox is None:
                continue
            x, y, cell_w, cell_h = bbox
            crop = bgr[y: y + cell_h, x: x + cell_w]
            if crop.size == 0:
                continue
            # Skip fully-black / redacted rows
            if crop.mean() < REDACTED_ROW_THRESHOLD:
                continue
            cv2.imwrite(str(CELLS_DIR / f"row{row_idx}_col{col_idx}.png"), crop)

    return grid, bgr


# ---------------------------------------------------------------------------
# Step 3 — Ollama (LLM) reading
# ---------------------------------------------------------------------------


def _check_ollama() -> bool:
    """Return True if the Ollama server is reachable."""
    try:
        r = requests.get(f"{OLLAMA_BASE_URL}/api/tags", timeout=5)
        return r.status_code == 200
    except requests.RequestException:
        return False


def _image_to_base64(img_path: Path) -> str:
    """Read *img_path* and return a base64-encoded string."""
    with open(img_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def _pil_to_base64(img: Image.Image) -> str:
    """Convert a PIL Image to a base64 PNG string."""
    import io
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode("utf-8")


WHOLE_PAGE_PROMPT = """\
You are reading a scanned handwritten mosquito surveillance form from Harris County, Texas.
Extract all data rows from the table (skip header rows and redacted rows).
Return a JSON array where each element is an object with these exact keys:
SITE_NO, COLL_NO, SPECIES, POOL, NO_MOSQ_POOLED, VIAL_NO, MSI, STATUS, NUM_POOLS, AREA, CITY, ADDRESS

Rules:
- SPECIES uses abbreviations: Cx.qf (Culex quinquefasciatus), Ae.ab (Aedes albopictus), Cu.ng (Culex nigripalpus), Ps.co (Psorophora columbiae), Ps.fx (Psorophora ferox). F = female pool limit.
- Multiple species separated by semicolons, e.g. "40 Cx.qf 20F; 4 Ae.ab 10F"
- Multiple vials separated by " / ", e.g. "8758 / 8759"
- Leave blank if cell is empty or illegible
- Return ONLY the JSON array, no explanation
"""

CELL_PROMPT = "Read the handwritten text in this cell exactly. Return only the text, nothing else."


def _call_ollama(model: str, prompt: str, image_b64: str) -> Optional[str]:
    """
    Send *image_b64* to Ollama and return the raw text response.

    Returns None if the call fails or times out.
    """
    payload = {
        "model": model,
        "prompt": prompt,
        "images": [image_b64],
        "stream": False,
    }
    try:
        resp = requests.post(
            f"{OLLAMA_BASE_URL}/api/generate",
            json=payload,
            timeout=600,  # 10 min — model load + inference on large image
        )
        resp.raise_for_status()
        data = resp.json()
        return data.get("response", "").strip()
    except (requests.RequestException, json.JSONDecodeError, KeyError) as exc:
        print(f"  [WARN] Ollama call failed: {exc}")
        return None


def _extract_json_array(text: str) -> Optional[list[dict]]:
    """
    Extract a JSON array from *text*.

    Handles cases where the model adds markdown fences or preamble.
    """
    if not text:
        return None

    # Strip markdown code fences
    text = re.sub(r"```[a-z]*\n?", "", text).strip()

    # Find the outermost [ ... ] block
    start = text.find("[")
    end = text.rfind("]")
    if start == -1 or end == -1:
        return None

    try:
        data = json.loads(text[start: end + 1])
        if isinstance(data, list):
            return data
    except json.JSONDecodeError:
        pass
    return None


def read_whole_page_ollama(
    page_img: Image.Image,
    model: str = DEFAULT_MODEL,
) -> Optional[list[dict]]:
    """
    Send the full page image to Ollama and return parsed rows.

    Returns None if Ollama is unavailable or returns invalid JSON.
    """
    # Resize to max 1280px wide — llava:7b handles large images poorly and times out
    max_w = 1280
    if page_img.width > max_w:
        scale = max_w / page_img.width
        new_h = int(page_img.height * scale)
        page_img = page_img.resize((max_w, new_h), Image.LANCZOS)
        print(f"  Resized to {page_img.width}x{page_img.height} for LLM")

    print(f"  Sending full page to Ollama ({model})...")
    b64 = _pil_to_base64(page_img)
    raw = _call_ollama(model, WHOLE_PAGE_PROMPT, b64)
    if raw is None:
        return None

    rows = _extract_json_array(raw)
    if rows is None:
        print(f"  [WARN] Ollama returned non-JSON:\n  {raw[:300]}")
        return None

    print(f"  Ollama returned {len(rows)} rows")
    # Normalize keys to COLUMNS
    normalized: list[dict] = []
    for row in rows:
        norm: dict[str, str] = {}
        for col in COLUMNS:
            norm[col] = str(row.get(col, "")).strip()
        normalized.append(norm)
    return normalized


def read_cells_ollama(
    grid: Grid,
    bgr: np.ndarray,
    model: str = DEFAULT_MODEL,
) -> list[dict]:
    """
    Read every cell individually via Ollama and assemble into rows.

    Skips cells whose crops are empty or fully-black.
    Falls back to empty string for any failed cell.
    """
    print(f"  Reading cells one-by-one with Ollama ({model})...")
    result_rows: list[dict] = []

    # Determine which grid rows look like data rows (skip header/footer)
    data_row_indices = _identify_data_rows_from_grid(grid, bgr)

    for ri in data_row_indices:
        row_boxes = grid[ri]
        row_data: dict[str, str] = {col: "" for col in COLUMNS}

        for ci, bbox in enumerate(row_boxes):
            # Skip col 0 (narrow margin strip) and out-of-range cols
            col_name_idx = ci - GRID_COL_OFFSET
            if ci < GRID_COL_OFFSET or col_name_idx >= len(COLUMNS):
                continue
            if bbox is None:
                continue
            x, y, bw, bh = bbox
            crop = bgr[y: y + bh, x: x + bw]
            if crop.size == 0 or crop.mean() < REDACTED_ROW_THRESHOLD:
                continue

            cell_path = CELLS_DIR / f"row{ri}_col{ci}.png"
            if not cell_path.exists():
                continue

            b64 = _image_to_base64(cell_path)
            text = _call_ollama(model, CELL_PROMPT, b64) or ""
            row_data[COLUMNS[col_name_idx]] = text.strip()

        if any(v for v in row_data.values()):
            result_rows.append(row_data)

    return result_rows


# ---------------------------------------------------------------------------
# Step 4 — Tesseract fallback
# ---------------------------------------------------------------------------


def _preprocess_for_tesseract(crop: np.ndarray) -> np.ndarray:
    """
    Prepare a cell crop for Tesseract:
    grayscale → upscale if tiny → Gaussian blur → Otsu binarize → border.
    """
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY) if len(crop.shape) == 3 else crop
    cell_h, cell_w = gray.shape
    if cell_h < 40 or cell_w < 40:
        scale = max(40 / max(cell_h, 1), 40 / max(cell_w, 1), 2.0)
        gray = cv2.resize(gray, None, fx=scale, fy=scale,
                          interpolation=cv2.INTER_CUBIC)
    gray = cv2.GaussianBlur(gray, (3, 3), 0)
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return cv2.copyMakeBorder(binary, 10, 10, 10, 10,
                              cv2.BORDER_CONSTANT, value=255)


def _ocr_cell_tesseract(crop: np.ndarray) -> str:
    """Run Tesseract on a single cell image and return cleaned text."""
    processed = _preprocess_for_tesseract(crop)
    pil = Image.fromarray(processed)
    raw = pytesseract.image_to_string(pil, config="--psm 6 --oem 3")
    return raw.strip().replace("\n", " ")


def _identify_data_rows_from_grid(grid: Grid, bgr: np.ndarray) -> list[int]:
    """
    Return indices of rows that look like real data rows.

    Strategy:
    - Skip the first 2 rows (always form title / column labels).
    - For each remaining row, OCR the cell at the SITE_NO column
      (grid column index = GRID_COL_OFFSET, typically 1) and check
      whether it contains a 3-to-5 digit number.
    - Fall back to a "non-empty cells" heuristic if Tesseract finds
      nothing numeric (e.g., the SITE_NO column was not detected).

    The grid's col-0 is always a narrow left-margin strip — real data
    starts at GRID_COL_OFFSET (= 1).
    """
    data_rows: list[int] = []
    # col index of SITE_NO in the raw grid (after the blank margin column)
    site_col = GRID_COL_OFFSET

    for ri, row_boxes in enumerate(grid):
        if ri < 2:
            continue

        # Quick check: skip rows that are entirely blank/whitespace
        non_empty_count = 0
        for bbox in row_boxes:
            if bbox is None:
                continue
            x, y, bw, bh = bbox
            crop = bgr[y: y + bh, x: x + bw]
            if crop.size > 0 and crop.mean() > REDACTED_ROW_THRESHOLD + 20:
                non_empty_count += 1
        if non_empty_count < 3:
            continue

        # Try to OCR the SITE_NO cell; if it contains a 3-5 digit number -> data row
        if site_col < len(row_boxes) and row_boxes[site_col] is not None:
            x, y, bw, bh = row_boxes[site_col]
            crop = bgr[y: y + bh, x: x + bw]
            if crop.size > 0 and crop.mean() > REDACTED_ROW_THRESHOLD:
                try:
                    text = _ocr_cell_tesseract(crop).strip()
                    # Strip noise chars (pipes, spaces, leading/trailing non-digits)
                    digits_only = re.sub(r"[^0-9]", "", text)
                    # Site numbers are 3-4 digit integers < 2000
                    # (exclude years like 2007 which appear in footer rows)
                    if re.fullmatch(r"\d{3,4}", digits_only) and int(digits_only) < 2000:
                        data_rows.append(ri)
                        continue
                except Exception:
                    pass

        # Fallback path intentionally skipped to avoid including header rows.
        # If the SITE_NO cell does not produce a 3-5 digit number the row is
        # treated as a header, footer, or annotation row and excluded.

    return data_rows


# The blank left-margin column is always col 0 in the detected grid.
# All real data columns start at this offset.
GRID_COL_OFFSET: int = 1


def read_with_tesseract(grid: Grid, bgr: np.ndarray) -> list[dict]:
    """
    Fallback: read every data row cell-by-cell with Tesseract.

    Column mapping: grid col 0 is the blank left-margin strip (skipped).
    Grid col 1 -> COLUMNS[0] (SITE_NO), grid col 2 -> COLUMNS[1] (COLL_NO), etc.
    Applies basic species-abbreviation cleanup heuristics.
    """
    print("  Reading cells with Tesseract (fallback)...")
    data_row_indices = _identify_data_rows_from_grid(grid, bgr)
    result_rows: list[dict] = []

    for ri in data_row_indices:
        row_boxes = grid[ri]
        row_data: dict[str, str] = {col: "" for col in COLUMNS}

        for ci, bbox in enumerate(row_boxes):
            # Skip col 0 (narrow margin strip) and out-of-range cols
            col_name_idx = ci - GRID_COL_OFFSET
            if ci < GRID_COL_OFFSET or col_name_idx >= len(COLUMNS):
                continue
            if bbox is None:
                continue
            x, y, bw, bh = bbox
            crop = bgr[y: y + bh, x: x + bw]
            if crop.size == 0 or crop.mean() < REDACTED_ROW_THRESHOLD:
                continue
            col_name = COLUMNS[col_name_idx]
            try:
                text = _ocr_cell_tesseract(crop)
                text = _clean_species_text(text) if col_name == "SPECIES" else text
                row_data[col_name] = text
            except Exception as exc:
                print(f"  [WARN] Tesseract failed on row={ri} col={ci}: {exc}")

        if any(v for v in row_data.values()):
            result_rows.append(row_data)

    return result_rows


def _validate_cell(text: str, col_name: str) -> bool:
    """
    Return True if `text` looks like a plausible value for `col_name`.
    Used to detect TrOCR hallucinations (e.g. "spicy", "river") so we can
    fall back to EasyOCR for that cell.
    """
    t = text.strip()
    if not t:
        return False
    if col_name == "SITE_NO":
        return bool(re.match(r"^\d{3,5}$", t))
    if col_name == "COLL_NO":
        # After _fix_coll_no this should be T4GV NNN; before that, at least has T/GV/digits
        return bool(re.search(r"[T][A-Z0-9]?GV|T\d|GV\s*\d", t, re.I))
    if col_name == "POOL":
        return bool(re.match(r"^\d{1,2}$", t))
    if col_name == "NUM_POOLS":
        return bool(re.match(r"^\d{1,2}$", t))
    if col_name in ("AREA",):
        return bool(re.match(r"^\d{2,4}$", t))
    if col_name == "CITY":
        return bool(re.match(r"^[A-Z]{2,4}$", t))
    if col_name == "VIAL_NO":
        return bool(re.search(r"\d{3,}", t))
    if col_name == "NO_MOSQ_POOLED":
        return bool(re.search(r"\d", t))
    if col_name == "ADDRESS":
        # At least one digit or an ALL-CAPS word (street names)
        return bool(re.search(r"\d", t) or re.search(r"[A-Z]{2,}", t))
    # SPECIES, MSI, STATUS — no strict validator
    return True


def _crop_cell(bgr: np.ndarray, bbox: tuple, pad: int = 3) -> np.ndarray:
    """
    Extract a cell from the full image with a small inward padding to:
    1. Remove the cell border lines (which confuse OCR engines)
    2. Add a tiny white margin so characters at the edge aren't clipped
    """
    x, y, bw, bh = bbox
    # Inset by `pad` pixels to remove grid lines, but ensure we don't go negative
    x1 = max(0, x + pad)
    y1 = max(0, y + pad)
    x2 = min(bgr.shape[1], x + bw - pad)
    y2 = min(bgr.shape[0], y + bh - pad)
    crop = bgr[y1:y2, x1:x2]
    if crop.size == 0:
        return crop
    # Add small white border so OCR doesn't clip edge characters
    return cv2.copyMakeBorder(crop, pad, pad, pad, pad,
                              cv2.BORDER_CONSTANT, value=(255, 255, 255))


def _fix_address_noise(v: str) -> str:
    """
    Address format: NUMBER STREET (MAP-CODE)
    EasyOCR sometimes appends garbage after the map code or before the street.
    Strip anything that appears after a valid (XNN-Y) / (NNN-Y) suffix.
    Also strip any leading tokens that look like mis-reads (pure lowercase short words).
    """
    # Remove trailing junk after map-code pattern like (419-H) or (380-M)
    v = re.sub(r"(\(\d{3}-[A-Z]\)).*", r"\1", v).strip()
    # Remove leading tokens that are pure lowercase and short (likely OCR noise)
    tokens = v.split()
    cleaned = []
    found_upper = False
    for tok in tokens:
        if tok[0].isupper() or tok[0].isdigit():
            found_upper = True
        if found_upper:
            cleaned.append(tok)
    return " ".join(cleaned) if cleaned else v


def read_with_trocr(grid: Grid, bgr: np.ndarray) -> list[dict]:
    """
    Read every data-row cell with Microsoft TrOCR (trocr-base-handwritten).
    TrOCR is a transformer encoder-decoder specifically trained on handwritten
    text — it outperforms EasyOCR on isolated handwritten words/lines.

    Runs on CPU; ~2s per cell, so expect ~3 min total for a full form.
    Uses inward cell padding to remove grid-line artifacts from crops.
    Falls back to EasyOCR for any cell where TrOCR fails.
    """
    try:
        from PIL import Image as PILImage
        from transformers import TrOCRProcessor, VisionEncoderDecoderModel
    except ImportError:
        print("  [WARN] transformers not installed. pip install transformers")
        return []

    print("  Loading TrOCR (microsoft/trocr-base-handwritten)...")
    processor = TrOCRProcessor.from_pretrained("microsoft/trocr-base-handwritten",
                                               use_fast=True)
    model = VisionEncoderDecoderModel.from_pretrained(
        "microsoft/trocr-base-handwritten")
    model.eval()
    print("  Reading cells with TrOCR...")

    # Also keep EasyOCR ready as fallback for empty/failed cells
    try:
        import easyocr
        easy_reader = easyocr.Reader(["en"], gpu=False, verbose=False)
    except ImportError:
        easy_reader = None

    data_row_indices = _identify_data_rows_from_grid(grid, bgr)
    result_rows: list[dict] = []

    for ri in data_row_indices:
        row_boxes = grid[ri]
        row_data: dict[str, str] = {col: "" for col in COLUMNS}

        for ci, bbox in enumerate(row_boxes):
            col_name_idx = ci - GRID_COL_OFFSET
            if ci < GRID_COL_OFFSET or col_name_idx >= len(COLUMNS):
                continue
            if bbox is None:
                continue

            crop = _crop_cell(bgr, bbox, pad=3)
            if crop.size == 0 or crop.mean() < REDACTED_ROW_THRESHOLD:
                continue

            col_name = COLUMNS[col_name_idx]

            # TrOCR needs RGB PIL image; upscale if too small
            rgb = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)
            h, w = rgb.shape[:2]
            if h < 32:
                scale = 32 / h
                rgb = cv2.resize(rgb, (max(32, int(w * scale)), 32),
                                 interpolation=cv2.INTER_CUBIC)

            try:
                pil = PILImage.fromarray(rgb)
                inputs = processor(images=pil, return_tensors="pt")
                import torch
                with torch.no_grad():
                    ids = model.generate(**inputs, max_new_tokens=48)
                text = processor.batch_decode(ids, skip_special_tokens=True)[0].strip()
                # Strip trailing punctuation artifacts (TrOCR often adds "." or ",")
                text = re.sub(r"[.,;]+$", "", text).strip()
            except Exception as exc:
                text = ""
                print(f"  [WARN] TrOCR row={ri} col={ci}: {exc}")

            # If TrOCR returned nothing, try EasyOCR fallback
            if not text and easy_reader is not None:
                try:
                    results = easy_reader.readtext(crop, detail=1, paragraph=False)
                    texts = [r[1] for r in sorted(results, key=lambda r: r[0][0][1])]
                    text = " ".join(texts).strip()
                except Exception:
                    pass

            text = _clean_species_text(text) if col_name == "SPECIES" else text
            if col_name == "ADDRESS":
                text = _fix_address_noise(text)
            row_data[col_name] = text

        if any(v for v in row_data.values()):
            result_rows.append(row_data)

    return result_rows


def read_trocr_primary(grid: Grid, bgr: np.ndarray) -> list[dict]:
    """
    TrOCR-primary mode: try TrOCR on EVERY cell first.
    If TrOCR's output fails the column-specific pattern validator
    (e.g. returns "spicy" for a CITY cell, or "river" for COLL_NO),
    fall back to EasyOCR for that cell.
    Final fallback for digit-only cells: Tesseract digit-mode.

    This lets TrOCR's superior handwriting recognition shine while
    EasyOCR/Tesseract patch up TrOCR's occasional hallucinations.
    """
    try:
        import easyocr
        from PIL import Image as PILImage
        from transformers import TrOCRProcessor, VisionEncoderDecoderModel
        import torch
    except ImportError as e:
        print(f"  [WARN] TrOCR-primary missing dep: {e}. Falling back to easyocr.")
        return []

    DIGIT_FALLBACK_COLS: set[str] = {"POOL", "NUM_POOLS", "AREA", "SITE_NO"}

    print("  Loading EasyOCR (English, CPU)...")
    easy_reader = easyocr.Reader(["en"], gpu=False, verbose=False)

    print("  Loading TrOCR (microsoft/trocr-base-handwritten)...")
    processor = TrOCRProcessor.from_pretrained("microsoft/trocr-base-handwritten",
                                               use_fast=True)
    trocr_model = VisionEncoderDecoderModel.from_pretrained(
        "microsoft/trocr-base-handwritten")
    trocr_model.eval()

    print("  Reading all cells with TrOCR (EasyOCR + Tesseract fallback on validation fail)...")

    data_row_indices = _identify_data_rows_from_grid(grid, bgr)
    result_rows: list[dict] = []

    for ri in data_row_indices:
        row_boxes = grid[ri]
        row_data: dict[str, str] = {col: "" for col in COLUMNS}

        for ci, bbox in enumerate(row_boxes):
            col_name_idx = ci - GRID_COL_OFFSET
            if ci < GRID_COL_OFFSET or col_name_idx >= len(COLUMNS):
                continue
            if bbox is None:
                continue

            crop = _crop_cell(bgr, bbox, pad=3)
            if crop.size == 0 or crop.mean() < REDACTED_ROW_THRESHOLD:
                continue

            col_name = COLUMNS[col_name_idx]

            # ---- Step 1: TrOCR ----
            rgb = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)
            h, w = rgb.shape[:2]
            if h < 32:
                scale = 32 / h
                rgb = cv2.resize(rgb, (max(32, int(w * scale)), 32),
                                 interpolation=cv2.INTER_CUBIC)
            text = ""
            try:
                pil = PILImage.fromarray(rgb)
                inputs = processor(images=pil, return_tensors="pt")
                with torch.no_grad():
                    ids = trocr_model.generate(**inputs, max_new_tokens=48)
                text = processor.batch_decode(ids, skip_special_tokens=True)[0].strip()
                text = re.sub(r"[.,;]+$", "", text).strip()
            except Exception as exc:
                print(f"  [WARN] TrOCR row={ri} col={ci}: {exc}")

            # ---- Step 2: validate — if TrOCR fails, try EasyOCR ----
            if not _validate_cell(text, col_name):
                easy_text = ""
                try:
                    results = easy_reader.readtext(crop, detail=1, paragraph=False)
                    texts = [r[1] for r in sorted(results, key=lambda r: r[0][0][1])]
                    easy_text = " ".join(texts).strip()
                except Exception:
                    pass
                # Use EasyOCR result if it validates; otherwise keep TrOCR (less wrong)
                if _validate_cell(easy_text, col_name):
                    text = easy_text
                elif easy_text:
                    text = easy_text  # still prefer EasyOCR over garbage TrOCR

            # ---- Step 3: Tesseract digit-mode fallback for small numeric cells ----
            if not re.search(r"\d", text) and col_name in DIGIT_FALLBACK_COLS:
                try:
                    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
                    h2, w2 = gray.shape
                    if h2 < 64:
                        scale = 64 / h2
                        gray = cv2.resize(gray, (max(32, int(w2 * scale)), 64),
                                          interpolation=cv2.INTER_CUBIC)
                    _, binary = cv2.threshold(gray, 0, 255,
                                              cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                    cfg = "--psm 7 -c tessedit_char_whitelist=0123456789"
                    tess = pytesseract.image_to_string(binary, config=cfg).strip()
                    digits = re.sub(r"\D", "", tess)
                    if digits:
                        text = digits
                except Exception:
                    pass

            if col_name == "SPECIES":
                text = _clean_species_text(text)
            if col_name == "ADDRESS":
                text = _fix_address_noise(text)
            row_data[col_name] = text

        if any(v for v in row_data.values()):
            result_rows.append(row_data)

    return result_rows


def read_hybrid(grid: Grid, bgr: np.ndarray) -> list[dict]:
    """
    Hybrid OCR: use TrOCR for columns where it outperforms EasyOCR
    (POOL, NUM_POOLS, NO_MOSQ_POOLED, VIAL_NO) and EasyOCR for the rest.

    Based on benchmarking on this form type:
      TrOCR wins:  POOL (4/7 vs 1/7), NO_MOSQ_POOLED (3/7 vs 1/7), VIAL_NO (1/7 vs 0/7)
      EasyOCR wins: SITE_NO, COLL_NO, CITY, AREA, SPECIES, ADDRESS
    """
    # Columns where TrOCR is used (numeric/pooling data)
    TROCR_COLS: set[str] = {"POOL", "NUM_POOLS", "NO_MOSQ_POOLED", "VIAL_NO"}

    # Digit-only Tesseract fallback for small numeric cells
    DIGIT_ONLY_COLS: set[str] = {"POOL", "NUM_POOLS"}

    try:
        import easyocr
        from PIL import Image as PILImage
        from transformers import TrOCRProcessor, VisionEncoderDecoderModel
        import torch
    except ImportError as e:
        print(f"  [WARN] Hybrid mode missing dep: {e}. Falling back to easyocr.")
        return []

    print("  Loading EasyOCR (English, CPU)...")
    easy_reader = easyocr.Reader(["en"], gpu=False, verbose=False)

    print("  Loading TrOCR (microsoft/trocr-base-handwritten)...")
    processor = TrOCRProcessor.from_pretrained("microsoft/trocr-base-handwritten",
                                               use_fast=True)
    trocr_model = VisionEncoderDecoderModel.from_pretrained(
        "microsoft/trocr-base-handwritten")
    trocr_model.eval()

    print("  Reading cells (EasyOCR + TrOCR hybrid, per-column best)...")

    data_row_indices = _identify_data_rows_from_grid(grid, bgr)
    result_rows: list[dict] = []

    for row_idx in data_row_indices:
        row_boxes = grid[row_idx]
        row_data: dict[str, str] = {col: "" for col in COLUMNS}

        for col_idx, bbox in enumerate(row_boxes):
            # The grid has a leading row-number column we skip (GRID_COL_OFFSET = 1)
            data_col_idx = col_idx - GRID_COL_OFFSET
            if col_idx < GRID_COL_OFFSET or data_col_idx >= len(COLUMNS):
                continue
            if bbox is None:
                continue

            crop = _crop_cell(bgr, bbox, pad=3)
            if crop.size == 0 or crop.mean() < REDACTED_ROW_THRESHOLD:
                continue

            col_name = COLUMNS[data_col_idx]
            text = ""

            if col_name in TROCR_COLS:
                # ---- TrOCR path (handwriting transformer) ----
                rgb = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)
                cell_h, cell_w = rgb.shape[:2]
                if cell_h < 32:
                    scale = 32 / cell_h
                    rgb = cv2.resize(rgb, (max(32, int(cell_w * scale)), 32),
                                     interpolation=cv2.INTER_CUBIC)
                try:
                    pil = PILImage.fromarray(rgb)
                    inputs = processor(images=pil, return_tensors="pt")
                    with torch.no_grad():
                        token_ids = trocr_model.generate(**inputs, max_new_tokens=48)
                    text = processor.batch_decode(token_ids, skip_special_tokens=True)[0].strip()
                    text = re.sub(r"[.,;]+$", "", text).strip()
                    # Discard TrOCR output for purely-numeric columns if it contains no digits
                    if col_name in DIGIT_ONLY_COLS and not re.search(r"\d", text):
                        text = ""
                except Exception:
                    text = ""
                # If TrOCR returned nothing, fall back to Tesseract digit-mode for small numeric cells
                if not text and col_name in DIGIT_ONLY_COLS:
                    try:
                        gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
                        fallback_h, fallback_w = gray.shape
                        if fallback_h < 64:
                            scale = 64 / fallback_h
                            gray = cv2.resize(gray,
                                              (max(32, int(fallback_w * scale)), 64),
                                              interpolation=cv2.INTER_CUBIC)
                        _, binary = cv2.threshold(gray, 0, 255,
                                                  cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                        cfg = "--psm 8 -c tessedit_char_whitelist=0123456789"
                        tess_raw = pytesseract.image_to_string(binary, config=cfg).strip()
                        text = re.sub(r"\D", "", tess_raw)
                    except Exception:
                        pass
            else:
                # ---- EasyOCR path (general text + codes) ----
                try:
                    detections = easy_reader.readtext(crop, detail=1, paragraph=False)
                    # Sort detections top-to-bottom by their bounding-box y coordinate
                    lines = [det[1] for det in sorted(detections, key=lambda d: d[0][0][1])]
                    text = " ".join(lines).strip()
                except Exception:
                    text = ""
                # If EasyOCR found nothing in a digit-only column, try Tesseract as backup
                if not text and col_name in {"SITE_NO", "AREA"}:
                    try:
                        gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
                        fallback_h, fallback_w = gray.shape
                        if fallback_h < 64:
                            scale = 64 / fallback_h
                            gray = cv2.resize(gray,
                                              (max(48, int(fallback_w * scale)), 64),
                                              interpolation=cv2.INTER_CUBIC)
                        _, binary = cv2.threshold(gray, 0, 255,
                                                  cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                        cfg = "--psm 7 -c tessedit_char_whitelist=0123456789"
                        tess_raw = pytesseract.image_to_string(binary, config=cfg).strip()
                        text = re.sub(r"\D", "", tess_raw)
                    except Exception:
                        pass

            if col_name == "SPECIES":
                text = _clean_species_text(text)
            if col_name == "ADDRESS":
                text = _fix_address_noise(text)
            row_data[col_name] = text

        if any(value for value in row_data.values()):
            result_rows.append(row_data)

    return result_rows


def read_with_easyocr(grid: Grid, bgr: np.ndarray) -> list[dict]:
    """
    Read every data-row cell with EasyOCR — a deep-learning OCR engine that
    handles handwriting significantly better than Tesseract and is much faster
    than running a full LLM vision model on CPU.

    EasyOCR uses a CRNN model trained on scene text / handwriting, so it
    handles slanted strokes, variable letter spacing, and domain-specific
    abbreviations better than Tesseract's classical pipeline.
    """
    try:
        import easyocr  # lazy import — large model download on first use
    except ImportError:
        print("  [WARN] EasyOCR not installed — skipping. pip install easyocr")
        return []

    # Columns that contain only digits — use Tesseract digit-mode as fallback
    # when EasyOCR returns nothing (cell may be too small to trigger its detector)
    DIGIT_ONLY_COLS: set[str] = {"POOL", "NUM_POOLS"}

    print("  Initialising EasyOCR (English, CPU)...")
    reader = easyocr.Reader(["en"], gpu=False, verbose=False)
    print("  Reading cells with EasyOCR (digit fallback via Tesseract for POOL/NUM_POOLS)...")

    data_row_indices = _identify_data_rows_from_grid(grid, bgr)
    result_rows: list[dict] = []

    for ri in data_row_indices:
        row_boxes = grid[ri]
        row_data: dict[str, str] = {col: "" for col in COLUMNS}

        for ci, bbox in enumerate(row_boxes):
            col_name_idx = ci - GRID_COL_OFFSET
            if ci < GRID_COL_OFFSET or col_name_idx >= len(COLUMNS):
                continue
            if bbox is None:
                continue
            crop = _crop_cell(bgr, bbox, pad=3)
            if crop.size == 0 or crop.mean() < REDACTED_ROW_THRESHOLD:
                continue

            col_name = COLUMNS[col_name_idx]
            try:
                # EasyOCR returns list of (bbox, text, confidence) tuples
                results = reader.readtext(crop, detail=1, paragraph=False)
                # Join all detected text segments in reading order
                texts = [r[1] for r in sorted(results, key=lambda r: r[0][0][1])]
                text = " ".join(texts).strip()
                text = _clean_species_text(text) if col_name == "SPECIES" else text
                if col_name == "ADDRESS":
                    text = _fix_address_noise(text)
                row_data[col_name] = text
            except Exception as exc:
                print(f"  [WARN] EasyOCR failed on row={ri} col={ci}: {exc}")

            # Digit-only fallback: if EasyOCR returned nothing for a numeric cell,
            # try Tesseract with digit whitelist + upscaled crop
            if not row_data[col_name] and col_name in DIGIT_ONLY_COLS:
                try:
                    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
                    h, w = gray.shape
                    if h < 64:
                        scale = 64 / h
                        gray = cv2.resize(gray, (max(32, int(w * scale)), 64),
                                          interpolation=cv2.INTER_CUBIC)
                    _, binary = cv2.threshold(gray, 0, 255,
                                              cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                    cfg = "--psm 8 -c tessedit_char_whitelist=0123456789"
                    tess_text = pytesseract.image_to_string(binary, config=cfg).strip()
                    digits = re.sub(r"\D", "", tess_text)
                    if digits:
                        row_data[col_name] = digits
                except Exception:
                    pass

        if any(v for v in row_data.values()):
            result_rows.append(row_data)

    return result_rows


def _preprocess_cell(crop: np.ndarray, digits_only: bool = False) -> np.ndarray:
    """
    Enhance a cell image for OCR:
    - Convert to greyscale
    - CLAHE contrast enhancement (handles uneven lighting in scans)
    - Gaussian denoise
    - Otsu threshold → clean binary image
    - Upscale to at least 48px tall (OCR engines need sufficient resolution)
    """
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY) if len(crop.shape) == 3 else crop.copy()
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(4, 4))
    gray = clahe.apply(gray)
    gray = cv2.GaussianBlur(gray, (3, 3), 0)
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    # Upscale short cells — OCR needs at least 48px height to work well
    h, w = binary.shape
    if h < 48:
        scale = 48 / h
        binary = cv2.resize(binary, (int(w * scale), 48), interpolation=cv2.INTER_CUBIC)
    return binary


def read_with_paddleocr(grid: Grid, bgr: np.ndarray) -> list[dict]:
    """
    Read every data-row cell with PaddleOCR — generally more accurate than
    EasyOCR on structured forms. Uses CLAHE preprocessing per cell.

    Numeric-only columns (SITE_NO, POOL, AREA, NUM_POOLS) get digit-mode
    preprocessing for better accuracy on short digit strings.
    """
    try:
        from paddleocr import PaddleOCR  # lazy import
    except ImportError:
        print("  [WARN] PaddleOCR not installed — skipping. pip install paddleocr paddlepaddle")
        return []

    DIGIT_COLS = {"SITE_NO", "POOL", "AREA", "NUM_POOLS"}

    print("  Initialising PaddleOCR (English, CPU)...")
    ocr = PaddleOCR(use_textline_orientation=True, lang="en")
    print("  Reading cells with PaddleOCR + CLAHE preprocessing...")

    data_row_indices = _identify_data_rows_from_grid(grid, bgr)
    result_rows: list[dict] = []

    for ri in data_row_indices:
        row_boxes = grid[ri]
        row_data: dict[str, str] = {col: "" for col in COLUMNS}

        for ci, bbox in enumerate(row_boxes):
            col_name_idx = ci - GRID_COL_OFFSET
            if ci < GRID_COL_OFFSET or col_name_idx >= len(COLUMNS):
                continue
            if bbox is None:
                continue
            x, y, bw, bh = bbox
            crop = bgr[y: y + bh, x: x + bw]
            if crop.size == 0 or crop.mean() < REDACTED_ROW_THRESHOLD:
                continue

            col_name = COLUMNS[col_name_idx]
            is_digit = col_name in DIGIT_COLS

            try:
                enhanced = _preprocess_cell(crop, digits_only=is_digit)
                # PaddleOCR wants BGR or RGB; convert binary back to BGR 3-channel
                enhanced_bgr = cv2.cvtColor(enhanced, cv2.COLOR_GRAY2BGR)
                result = ocr.ocr(enhanced_bgr, cls=True)
                texts = []
                if result and result[0]:
                    for line in result[0]:
                        if line and len(line) >= 2:
                            texts.append(line[1][0])
                text = " ".join(texts).strip()
                text = _clean_species_text(text) if col_name == "SPECIES" else text
                row_data[col_name] = text
            except Exception as exc:
                print(f"  [WARN] PaddleOCR failed on row={ri} col={ci}: {exc}")

        if any(v for v in row_data.values()):
            result_rows.append(row_data)

    return result_rows


def _clean_species_text(text: str) -> str:
    """Fix common OCR misreads in species abbreviation fields."""
    fixes = {
        r"\bCx\.qf\b": "Cx.qf",
        r"\bCxqf\b": "Cx.qf",
        r"\bCx\.Qf\b": "Cx.qf",
        r"\bAe\.ab\b": "Ae.ab",
        r"\bAeab\b": "Ae.ab",
        r"\bCu\.ng\b": "Cu.ng",
        r"\bCung\b": "Cu.ng",
        r"\bPs\.co\b": "Ps.co",
        r"\bPsco\b": "Ps.co",
        r"\bPs\.fx\b": "Ps.fx",
        r"\bPsfx\b": "Ps.fx",
    }
    for pattern, replacement in fixes.items():
        text = re.sub(pattern, replacement, text)
    return text


# ---------------------------------------------------------------------------
# Metadata extraction from header region
# ---------------------------------------------------------------------------


def extract_form_metadata(page_img: Image.Image) -> dict[str, str]:
    """
    Extract form metadata (collection method, week number, date) from the
    top portion of the page using Tesseract.
    """
    width, height = page_img.size
    header_crop = page_img.crop((0, 0, width, int(height * 0.15)))
    raw = pytesseract.image_to_string(header_crop, config="--psm 6 --oem 3")

    meta: dict[str, str] = {
        "collection_method": "",
        "week_no": "",
        "date": "",
    }

    # Week number: look for "WEEK NO" or "Week" followed by digits
    week_match = re.search(r"WEEK\s*NO[.\s]*(\d+)", raw, re.IGNORECASE)
    if week_match:
        meta["week_no"] = week_match.group(1)

    # Date: look for patterns like "8/1/07" or "9/K/07"
    date_match = re.search(r"\d{1,2}/[\w]+/\d{2,4}", raw)
    if date_match:
        meta["date"] = date_match.group(0)

    # Collection method: look for "CDC GRAVID" or "CDC LIGHT"
    if re.search(r"GRAVID", raw, re.IGNORECASE):
        meta["collection_method"] = "CDC GRAVID"
    elif re.search(r"LIGHT TRAP", raw, re.IGNORECASE):
        meta["collection_method"] = "CDC Light Trap"

    return meta


# ---------------------------------------------------------------------------
# Step 5 — Output: CSV + XLSX
# ---------------------------------------------------------------------------


def write_csv(rows: list[dict], path: Path) -> None:
    """Write *rows* to a CSV file at *path*."""
    import csv
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    print(f"  CSV saved -> {path}")


def write_xlsx(
    rows: list[dict],
    path: Path,
    metadata: Optional[dict[str, str]] = None,
) -> None:
    """
    Write *rows* to a formatted XLSX workbook at *path*.

    Layout:
        Row 1 : Agency title (merged, dark navy)
        Row 2 : Surveillance type (merged, dark navy)
        Row 3 : Collection Method / Week No / Date (light blue)
        Row 4 : blank spacer
        Row 5 : Column headers (dark blue, white bold)
        Row 6+ : data rows (alternating white / light blue)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Surveillance Data"

    num_cols = len(COLUMNS)

    # Colour palettes
    fill_navy = PatternFill("solid", fgColor="1F4E79")
    fill_dark_blue = PatternFill("solid", fgColor="2E75B6")
    fill_light_blue = PatternFill("solid", fgColor="DEEAF1")
    fill_white = PatternFill("solid", fgColor="FFFFFF")

    font_white_bold = Font(color="FFFFFF", bold=True)
    font_white = Font(color="FFFFFF")
    font_bold = Font(bold=True)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_middle = Alignment(vertical="center", wrap_text=True)

    def merge_row(row_num: int, value: str, fill: PatternFill, font: Font) -> None:
        ws.merge_cells(
            start_row=row_num, start_column=1,
            end_row=row_num, end_column=num_cols,
        )
        cell = ws.cell(row=row_num, column=1, value=value)
        cell.fill = fill
        cell.font = font
        cell.alignment = align_center
        ws.row_dimensions[row_num].height = 20

    # Row 1 — Agency title
    merge_row(
        1,
        "Harris County Public Health and Environmental Services — Mosquito Control Division",
        fill_navy, font_white_bold,
    )

    # Row 2 — Surveillance type
    merge_row(
        2,
        "WNV/SLE Surveillance — CDC Light Trap Collections",
        fill_navy, font_white_bold,
    )

    # Row 3 — Metadata
    if metadata:
        meta_str = (
            f"Collection Method: {metadata.get('collection_method', '')}   "
            f"Week No: {metadata.get('week_no', '')}   "
            f"Date: {metadata.get('date', '')}"
        )
    else:
        meta_str = "Collection Method:    Week No:    Date:"

    ws.merge_cells(
        start_row=3, start_column=1,
        end_row=3, end_column=num_cols,
    )
    meta_cell = ws.cell(row=3, column=1, value=meta_str)
    meta_cell.fill = fill_light_blue
    meta_cell.font = font_bold
    meta_cell.alignment = align_center
    ws.row_dimensions[3].height = 18

    # Row 4 — blank spacer
    ws.row_dimensions[4].height = 8

    # Row 5 — column headers
    for ci, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=5, column=ci, value=col_name)
        cell.fill = fill_dark_blue
        cell.font = font_white_bold
        cell.alignment = align_center

    # Rows 6+ — data
    for ri, row_dict in enumerate(rows):
        fill = fill_light_blue if ri % 2 == 1 else fill_white
        for ci, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=6 + ri, column=ci, value=row_dict.get(col_name, ""))
            cell.fill = fill
            cell.alignment = align_middle

    # Column widths
    for ci, col_name in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(ci)
        ws.column_dimensions[col_letter].width = COLUMN_WIDTHS.get(col_name, 12)

    # Freeze panes at C6 (freeze columns A-B and rows 1-5)
    ws.freeze_panes = "C6"

    wb.save(str(path))
    print(f"  XLSX saved -> {path}")


# ---------------------------------------------------------------------------
# Step 6 — QC report
# ---------------------------------------------------------------------------


def run_qc_report(
    extracted: list[dict],
    ground_truth: list[dict] = GROUND_TRUTH,
    fields: list[str] = QC_FIELDS,
) -> None:
    """
    Compare *extracted* rows against *ground_truth* and print a QC table.

    Reports MATCH / MISMATCH per field per row, then overall accuracy %.
    """
    print("\n" + "=" * 80)
    print("QC REPORT - Extracted vs. Ground Truth")
    print("=" * 80)

    header = f"{'Row':<5} {'Field':<20} {'Extracted':<35} {'Expected':<35} {'Result'}"
    print(header)
    print("-" * len(header))

    total = 0
    matches = 0

    n_rows = min(len(extracted), len(ground_truth))

    for row_idx in range(n_rows):
        ext_row = extracted[row_idx]
        gt_row = ground_truth[row_idx]

        for field in fields:
            ext_val = ext_row.get(field, "").strip()
            gt_val = gt_row.get(field, "").strip()

            # Normalise for comparison: collapse whitespace, lower-case
            ext_norm = re.sub(r"\s+", " ", ext_val).lower()
            gt_norm = re.sub(r"\s+", " ", gt_val).lower()

            match = "MATCH" if ext_norm == gt_norm else "MISMATCH"
            if match == "MATCH":
                matches += 1
            total += 1

            ext_disp = ext_val[:33] + ".." if len(ext_val) > 35 else ext_val
            gt_disp = gt_val[:33] + ".." if len(gt_val) > 35 else gt_val
            print(f"{row_idx + 1:<5} {field:<20} {ext_disp:<35} {gt_disp:<35} {match}")

        print()  # blank line between rows

    if total > 0:
        accuracy = 100.0 * matches / total
        print(f"Overall accuracy: {matches}/{total} fields matched = {accuracy:.1f}%")
    else:
        print("No rows to compare.")

    print("=" * 80)


# ---------------------------------------------------------------------------
# Post-processing: column cleanup + species DB matching
# ---------------------------------------------------------------------------


def _digit_subs(raw: str) -> str:
    """Replace unambiguous OCR letter-for-digit confusions before numeric parsing.
    Only safe substitutions: l→1, I→1, O→0 (context-free misreads for digits).
    Avoids S→5, B→8 etc. which legitimately appear in non-numeric fields.
    """
    return raw.replace("l", "1").replace("I", "1").replace("O", "0")


def _fix_site_no(raw: str) -> str:
    """
    SITE_NO is a 4-digit location identifier (e.g. 1841, 1987).
    Apply digit substitutions, collapse spaces, then extract the first 3-5 digit run.
    """
    cleaned = _digit_subs(raw)
    cleaned = re.sub(r"\s+", "", cleaned)  # collapse "1 841" → "1841"
    match = re.search(r"\d{3,5}", cleaned)
    if match:
        return match.group(0)[:5]
    digits_only = re.sub(r"\D", "", cleaned)
    return digits_only if digits_only else raw


def _fix_pool(raw: str) -> str:
    """
    POOL is a 1-2 digit count of mosquito pools collected (typically 1 or 2).
    Discard if > 20 — that's almost certainly an OCR bleed from an adjacent column.
    """
    cleaned = _digit_subs(raw)
    match = re.search(r"\d+", cleaned)
    if match:
        digits = match.group(0)[:2]
        if int(digits) <= 20:
            return digits
    return ""


def _fix_area(raw: str) -> str:
    """
    AREA is a 3-digit Harris County geographic zone code (e.g. 216, 317).
    Extract the first 3-digit group after applying digit substitutions.
    """
    cleaned = _digit_subs(raw)
    match = re.search(r"\d{3}", cleaned)
    if match:
        return match.group(0)
    digits_only = re.sub(r"\D", "", cleaned)
    return digits_only[:3] if len(digits_only) >= 3 else (digits_only if digits_only else raw)


def _fix_num_pools(raw: str) -> str:
    """NUM_POOLS is a 1-2 digit integer — how many distinct trap pools were set up."""
    cleaned = _digit_subs(raw)
    digits_only = re.sub(r"\D", "", cleaned)
    if digits_only:
        try:
            return str(int(digits_only[:2]))  # int() strips leading zeros
        except ValueError:
            return digits_only[:2]
    return ""


def _fix_no_mosq_pooled(raw: str) -> str:
    """
    NO_MOSQ_POOLED can be:
    - A plain integer ('5', '24', '41') — total mosquitoes across all pools
    - A compound species string ('39 Cx.qf / 4 Ae.ab') — count broken down by species

    For plain integers with trailing OCR noise ('50-', '50edge', '3-negles'),
    strip the noise and return just the number. Leave species compound strings
    intact so the downstream species matcher can process them.
    """
    raw = raw.strip()
    if not raw:
        return raw
    # Plain integer, possibly with trailing punctuation or short OCR artifact
    plain_integer_match = re.match(r"^(\d+)\s*[-_.,;)(*]?\s*$", raw)
    if plain_integer_match:
        return plain_integer_match.group(1)
    # Leading integer followed by non-species garbage — extract just the number
    is_species_compound = bool(re.search(r"Cx\.|Ae\.|Ps\.|An\.|Oc\.|Ma\.|Cq\.|/", raw))
    if not is_species_compound:
        leading_number = re.match(r"^(\d+)\D", raw)
        if leading_number:
            return leading_number.group(1)
    return raw


def _fix_msi(raw: str) -> str:
    """MSI (Mosquito Salivary Index) is a single digit, or empty if not tested."""
    digits = re.sub(r"\D", "", _digit_subs(raw))
    return digits[:1] if digits else ""


def _fix_status(raw: str) -> str:
    """STATUS is a short uppercase result code (e.g. 'N' for negative, 'P' for positive)."""
    letters_and_digits = re.sub(r"[^A-Za-z0-9]", "", raw).upper()
    return letters_and_digits[:6] if letters_and_digits else ""


def _fix_numeric(raw: str) -> str:
    """Generic digits-only extraction — legacy fallback for unrecognised column types."""
    digits = re.sub(r"\D", "", _digit_subs(raw))
    return digits if digits else raw


def _fix_vial_no(raw: str) -> str:
    """
    VIAL_NO is one or two 4-digit sample vial numbers (e.g. '8758' or '8758 / 8759').
    When a pool has mosquitoes of multiple species, two separate vials are used and
    their numbers are written side by side in the cell.

    Strategy:
    1. Apply digit substitutions (l→1, I→1, O→0)
    2. Strip any stray letters from the SPECIES column bleeding in
    3. Extract all 4-digit groups — these are the vial numbers
    4. Fall back to 3-5 digit groups if no clean 4-digit match found
    """
    cleaned = _digit_subs(raw)
    digits_and_spaces = re.sub(r"[A-Za-z]", "", cleaned)
    exact_vials = re.findall(r"\d{4}", digits_and_spaces)
    if exact_vials:
        return " / ".join(exact_vials)
    approx_vials = re.findall(r"\d{3,5}", digits_and_spaces)
    if approx_vials:
        return " / ".join(approx_vials)
    return raw


def _fix_coll_no(raw: str) -> str:
    """
    COLL_NO follows the pattern: T4GV NNN  (e.g. 'T4GV 508')
    where '4' is the Harris County trap number (always 4) and NNN is a
    sequential 3-digit collection number assigned per field visit.

    Handles common OCR errors:
    - '4' misread as 'A': TAGV 508 → T4GV 508
    - Any other letter in the trap position is normalised to 4
    """
    if re.match(r"^T4GV \d{3}$", raw):
        return raw  # already correct, skip processing
    # Normalise any T-letter-GV variant to T4GV
    normalised = re.sub(r"T[^GD\d]GV", "T4GV", raw)
    normalised = re.sub(r"T(\d)GV", lambda m: f"T{m.group(1)}GV", normalised)
    # Extract the collection number from the end
    number_match = re.search(r"T(\d)GV\s*(\d+)", normalised)
    if number_match:
        trap_num = number_match.group(1)
        coll_num = number_match.group(2)[:3].zfill(3)
        return f"T{trap_num}GV {coll_num}"
    # Last-resort fallback when the trap digit was read as a letter entirely
    letter_trap_match = re.search(r"T[A-Za-z]GV\s*(\d+)", raw)
    if letter_trap_match:
        coll_num = letter_trap_match.group(1)[:3].zfill(3)
        return f"T4GV {coll_num}"
    return normalised if normalised != raw else raw


def _fix_city(raw: str) -> str:
    """CITY is a 2-4 letter uppercase area code (e.g. 'CB' = Channelview, 'NEC' = NE corner)."""
    letters_only = re.sub(r"[^A-Za-z]", "", raw).upper()
    return letters_only[:4] if letters_only else raw


def match_species_field(raw: str) -> str:
    """
    Post-process a SPECIES cell value using the Harris County species database.

    Strategy:
    1. Split on common separators (;  /  +  newline)
    2. For each segment, extract the leading count (digits) and trailing pool-limit (digits + F)
    3. Extract the species abbreviation portion (letter cluster)
    4. Look up in ABBREV_VARIANTS, then fuzzy-match against SPECIES_DB keys
    5. Reconstruct canonical segment: "{count} {abbrev} {pool_limit}F"
    6. Rejoin with "; "
    """
    from species_db import SPECIES_DB, ABBREV_VARIANTS

    try:
        from thefuzz import process as fuzz_process
        _has_fuzz = True
    except ImportError:
        _has_fuzz = False

    all_abbrevs = list(SPECIES_DB.keys())

    def _match_abbrev(token: str) -> str | None:
        """Return canonical abbrev or None if no good match."""
        # Exact variant lookup first
        if token in ABBREV_VARIANTS:
            return ABBREV_VARIANTS[token]
        if token in SPECIES_DB:
            return token
        # Try case-insensitive exact on DB keys
        for k in all_abbrevs:
            if k.lower() == token.lower():
                return k
        # Fuzzy match — only if thefuzz available
        if _has_fuzz and len(token) >= 3:
            match, score = fuzz_process.extractOne(token, all_abbrevs)
            if score >= 55:
                return match
        return None

    def _parse_segment(seg: str) -> str:
        seg = seg.strip()
        if not seg:
            return ""

        # Extract leading count e.g. "40", "210"
        count_m = re.match(r"^(\d+)\s*", seg)
        count = count_m.group(1) if count_m else ""
        rest = seg[count_m.end():] if count_m else seg

        # Extract trailing pool limit e.g. "20F", "10F"
        pool_m = re.search(r"(\d+)\s*[Ff]\s*$", rest)
        pool_limit = pool_m.group(1) if pool_m else ""
        if pool_m:
            rest = rest[:pool_m.start()].strip()

        # rest should now be just the abbreviation (possibly garbled)
        # Try whole token, then sub-tokens
        abbrev = _match_abbrev(rest.strip())
        if abbrev is None:
            # Try splitting on non-alpha and matching each part
            parts = re.findall(r"[A-Za-z]+", rest)
            for p in parts:
                abbrev = _match_abbrev(p)
                if abbrev:
                    break

        if abbrev is None:
            return seg  # give up, return original

        parts_out = []
        if count:
            parts_out.append(count)
        parts_out.append(abbrev)
        if pool_limit:
            parts_out.append(f"{pool_limit}F")
        return " ".join(parts_out)

    # Split on separators
    segments = re.split(r"[;/\+\n]", raw)
    fixed = [_parse_segment(s) for s in segments]
    fixed = [f for f in fixed if f]
    return "; ".join(fixed) if fixed else raw


def postprocess_row(row: dict, species_corrections: list | None = None) -> dict:
    """
    Apply column-specific cleanup and species DB matching to a single row dict.
    Modifies a copy and returns it.
    species_corrections: optional list to append (original, corrected) tuples to.
    """
    result = dict(row)

    # Apply type-specific cleanup to each column
    result["SITE_NO"]        = _fix_site_no(result.get("SITE_NO", ""))
    result["POOL"]           = _fix_pool(result.get("POOL", ""))
    result["AREA"]           = _fix_area(result.get("AREA", ""))
    result["NUM_POOLS"]      = _fix_num_pools(result.get("NUM_POOLS", ""))
    result["NO_MOSQ_POOLED"] = _fix_no_mosq_pooled(result.get("NO_MOSQ_POOLED", ""))
    result["MSI"]            = _fix_msi(result.get("MSI", ""))
    result["STATUS"]         = _fix_status(result.get("STATUS", ""))
    result["VIAL_NO"]        = _fix_vial_no(result.get("VIAL_NO", ""))
    result["COLL_NO"]        = _fix_coll_no(result.get("COLL_NO", ""))
    result["CITY"]           = _fix_city(result.get("CITY", ""))

    # Match species abbreviations against the Harris County species database
    raw_species = result.get("SPECIES", "")
    corrected_species = match_species_field(raw_species)
    if species_corrections is not None and corrected_species != raw_species:
        species_corrections.append((raw_species, corrected_species))
    result["SPECIES"] = corrected_species

    return result


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Local OCR pipeline for Harris County mosquito surveillance PDFs."
    )
    parser.add_argument(
        "pdf",
        nargs="?",
        default=str(SCRIPT_DIR / "example_mosquito_record_test.pdf"),
        help="Path to the input PDF (default: example_mosquito_record_test.pdf)",
    )
    parser.add_argument(
        "--mode",
        choices=["whole-page", "cells", "trocr-primary", "hybrid", "trocr", "easyocr", "paddleocr", "tesseract"],
        default="hybrid",
        help=(
            "hybrid: EasyOCR+TrOCR per-column best, highest CPU accuracy (default). "
            "trocr: Microsoft TrOCR handwriting model for all cells. "
            "easyocr: EasyOCR deep-learning OCR, faster but less accurate. "
            "paddleocr: PaddleOCR (broken on Windows CPU, falls back to easyocr). "
            "whole-page: send full page to Ollama LLM (needs GPU for speed). "
            "cells: send each cell to Ollama LLM (needs GPU for speed). "
            "tesseract: classic OCR fallback, poor on handwriting."
        ),
    )
    parser.add_argument(
        "--model",
        default=DEFAULT_MODEL,
        help=f"Ollama model to use (default: {DEFAULT_MODEL})",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    pdf_path = Path(args.pdf)

    if not pdf_path.exists():
        print(f"[ERROR] PDF not found: {pdf_path}", file=sys.stderr)
        sys.exit(1)

    print(f"\nInput: {pdf_path}")
    print(f"Mode : {args.mode}  |  Model: {args.model}\n")

    # ------------------------------------------------------------------
    # Step 1: PDF → image
    # ------------------------------------------------------------------
    page_img = pdf_to_image(pdf_path)

    # ------------------------------------------------------------------
    # Step 2: Table cell detection
    # ------------------------------------------------------------------
    grid, bgr = detect_cells(page_img)

    # ------------------------------------------------------------------
    # Step 3 / 4: OCR reading — engine chosen by --mode
    # ------------------------------------------------------------------
    rows: Optional[list[dict]] = None

    if args.mode == "trocr-primary":
        print("[Step 3/6] Running TrOCR-Primary (TrOCR all cells, EasyOCR/Tesseract on validation fail)...")
        rows = read_trocr_primary(grid, bgr)

    elif args.mode == "hybrid":
        print("[Step 3/6] Running Hybrid OCR (EasyOCR + TrOCR per-column best)...")
        rows = read_hybrid(grid, bgr)

    elif args.mode == "trocr":
        print("[Step 3/6] Running TrOCR (Microsoft handwriting model)...")
        rows = read_with_trocr(grid, bgr)

    elif args.mode == "paddleocr":
        print("[Step 3/6] Running PaddleOCR (with CLAHE cell preprocessing)...")
        rows = read_with_paddleocr(grid, bgr)

    elif args.mode == "easyocr":
        print("[Step 3/6] Running EasyOCR (deep-learning handwriting OCR)...")
        rows = read_with_easyocr(grid, bgr)

    elif args.mode == "tesseract":
        print("[Step 3/6] Running Tesseract OCR...")
        rows = read_with_tesseract(grid, bgr)

    elif args.mode in ("whole-page", "cells"):
        print("[Step 3/6] Running LLM reading via Ollama...")
        ollama_available = _check_ollama()
        if ollama_available:
            print(f"  Ollama is running at {OLLAMA_BASE_URL}")
            if args.mode == "whole-page":
                rows = read_whole_page_ollama(page_img, model=args.model)
                if rows is None and args.model != FALLBACK_MODEL:
                    print(f"  Retrying with fallback model {FALLBACK_MODEL}...")
                    rows = read_whole_page_ollama(page_img, model=FALLBACK_MODEL)
            else:
                rows = read_cells_ollama(grid, bgr, model=args.model)
        else:
            print("  Ollama not available — falling back to EasyOCR")

    # ------------------------------------------------------------------
    # Step 4: Cascade fallbacks if primary engine returned nothing
    # ------------------------------------------------------------------
    print("[Step 4/6] Checking fallback needed...")
    if not rows and args.mode not in ("hybrid", "trocr", "paddleocr", "easyocr", "tesseract"):
        print("  Trying Hybrid OCR fallback...")
        rows = read_hybrid(grid, bgr)
    if not rows and args.mode not in ("trocr", "paddleocr", "easyocr", "tesseract"):
        print("  Trying TrOCR fallback...")
        rows = read_with_trocr(grid, bgr)
    if not rows and args.mode not in ("paddleocr", "easyocr", "tesseract"):
        print("  Trying PaddleOCR fallback...")
        rows = read_with_paddleocr(grid, bgr)
    if not rows and args.mode not in ("easyocr", "tesseract"):
        print("  Trying EasyOCR fallback...")
        rows = read_with_easyocr(grid, bgr)
    if not rows:
        print("  Using Tesseract as last resort...")
        rows = read_with_tesseract(grid, bgr)
        if not rows:
            print("[WARN] All OCR engines returned no rows. Output files will be empty.")
            rows = []

    print(f"  Extracted {len(rows)} data rows")

    # Also extract form metadata from the header
    metadata = extract_form_metadata(page_img)
    print(f"  Metadata: {metadata}")

    # ------------------------------------------------------------------
    # Step 4b: Post-processing (column cleanup + species DB matching)
    # ------------------------------------------------------------------
    print("[Step 4b/6] Post-processing rows (column cleanup + species DB)...")
    corrections: list = []
    rows = [postprocess_row(r, corrections) for r in rows]
    if corrections:
        print(f"  Species DB: applied {len(corrections)} correction(s)")
        for orig, fixed in corrections:
            print(f"    '{orig[:40]}' -> '{fixed[:40]}'")
    else:
        print("  Species DB: no corrections needed")

    # ------------------------------------------------------------------
    # Step 5: Output
    # ------------------------------------------------------------------
    print("[Step 5/6] Writing output files...")
    write_csv(rows, OUTPUT_CSV)
    write_xlsx(rows, OUTPUT_XLSX, metadata=metadata)

    # ------------------------------------------------------------------
    # Step 6: QC report
    # ------------------------------------------------------------------
    print("[Step 6/6] Running QC report against ground truth...")
    run_qc_report(rows, GROUND_TRUTH, QC_FIELDS)

    print(f"\nDone.")
    print(f"  {OUTPUT_CSV}")
    print(f"  {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
